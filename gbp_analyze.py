#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gbp_analyze.py — ניתוח ייצוא Google Business Profile.

למה זה קיים: לחברת שירות עם עשרה סניפים, הכרטיס העסקי הוא מקור ליד ישיר —
שיחות טלפון, בקשות ניווט, וחיפושים מקומיים. עד 2026-08-10 לא היה לנו עליו
שום נתון. GSC אומרת מי הגיע לאתר; GBP אומר מי התקשר בלי להיכנס לאתר בכלל.

GBP API דורש אישור ידני מגוגל שלוקח ימים עד שבועות, ולכן מתחילים מייצוא.

איך מייצאים (פעם בחודש, כ-3 דקות):
  1. business.google.com → בחר פרופיל
  2. Performance (או "ביצועים")
  3. בחר טווח תאריכים — מומלץ 6 חודשים
  4. Download / הורדה → CSV
  5. שמור בשם: gbp_{site}_{branch}.csv  (למשל gbp_csb_lod.csv)
  6. העלה את כל הקבצים לתיקייה אחת והרץ:
       python3 gbp_analyze.py --dir /path/to/exports

הסקריפט מזהה את מבנה הקובץ לבד, כי גוגל משנה כותרות בין גרסאות ובין שפות.
"""
import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

OUT = Path("cats")

# גוגל משנה שמות עמודות בין גרסאות ובין עברית לאנגלית.
# מיפוי לפי מילות מפתch ולא לפי שם מדויק.
FIELD_MAP = {
    "calls": ["call", "שיחות", "טלפון", "phone"],
    "directions": ["direction", "ניווט", "מסלול", "route"],
    "website": ["website", "אתר", "click"],
    "views_search": ["search", "חיפוש"],
    "views_maps": ["maps", "מפות"],
    "bookings": ["booking", "הזמנ"],
    "messages": ["message", "הודע"],
}


# פורמט הייצוא בפועל (אומת 2026-08-10): שורה אחת לפרופיל, 16 עמודות,
# שתי שורות כותרת. גוגל הסירה את כפתור ה-CSV מהממשק החדש; הייצוא
# הזמין הוא xlsx מ-Business Profile Manager.
COLS = {
    "store_code": "Store code",
    "name": "Business name",
    "address": "Address",
    "search_mobile": "Google Search - Mobile",
    "search_desktop": "Google Search - Desktop",
    "maps_mobile": "Google Maps - Mobile",
    "maps_desktop": "Google Maps - Desktop",
    "calls": "Calls",
    "messages": "Messages",
    "bookings": "Bookings",
    "directions": "Directions",
    "website": "Website clicks",
}


def num(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def parse_file(path):
    """מחזיר רשומה לכל שורת פרופיל בקובץ."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []
    header = [str(c).strip() if c else "" for c in rows[0]]
    idx = {k: header.index(v) for k, v in COLS.items() if v in header}
    if "calls" not in idx:
        return []

    out = []
    for r in rows[1:]:
        if not r or not any(r):
            continue
        name = str(r[idx["name"]] or "").strip() if "name" in idx else ""
        if not name or name.startswith("Number of people"):
            continue
        rec = {k: (num(r[i]) if k not in ("store_code", "name", "address")
                   else str(r[i] or "").strip())
               for k, i in idx.items() if i < len(r)}
        rec["views"] = (rec.get("search_mobile", 0) + rec.get("search_desktop", 0)
                        + rec.get("maps_mobile", 0) + rec.get("maps_desktop", 0))
        rec["actions"] = (rec.get("calls", 0) + rec.get("directions", 0)
                          + rec.get("website", 0) + rec.get("messages", 0)
                          + rec.get("bookings", 0))
        rec["file"] = Path(path).name
        # אחוז המובייל — משפיע ישירות על מבנה התוכן
        mob = rec.get("search_mobile", 0) + rec.get("maps_mobile", 0)
        rec["mobile_share"] = round(100 * mob / rec["views"], 1) if rec["views"] else 0
        rec["action_rate"] = round(100 * rec["actions"] / rec["views"], 2) if rec["views"] else 0
        rec["call_rate"] = round(100 * rec.get("calls", 0) / rec["views"], 2) if rec["views"] else 0
        out.append(rec)
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", required=True)
    a = ap.parse_args()
    d = Path(a.dir)
    files = sorted(list(d.glob("*.xlsx")) + list(d.glob("*.xls")))
    if not files:
        print(f"❌ אין קבצי xlsx ב-{d}", file=sys.stderr)
        return 1

    profiles, failed = [], []
    for f in files:
        try:
            recs = parse_file(f)
        except Exception as e:
            failed.append((f.name, str(e)[:80]))
            continue
        if recs:
            profiles.extend(recs)
        else:
            failed.append((f.name, "לא זוהו עמודות מוכרות"))

    if not profiles:
        print("❌ אף פרופיל לא נותח", file=sys.stderr)
        for n, e in failed:
            print(f"   {n}: {e}", file=sys.stderr)
        return 1

    profiles.sort(key=lambda r: -r["views"])
    tot = {k: sum(r.get(k, 0) for r in profiles)
           for k in ("views", "calls", "directions", "website", "actions")}

    L = ["# Google Business Profile — ביצועים", "",
         f"נוצר: {date.today()} | {len(profiles)} פרופילים", "",
         "**מה זה מוסיף:** GSC מודדת מי הגיע לאתר. כאן רואים מי **התקשר**, ",
         "ביקש ניווט או חיפש את הסניף — לידים שאינם עוברים דרך האתר כלל.", "",
         "| פרופיל | צפיות | שיחות | ניווט | לאתר | % פעולה | % שיחה | % מובייל |",
         "|---|---|---|---|---|---|---|---|"]
    for r in profiles:
        L.append(f"| {r.get('name','')[:30]} | {r['views']:,} | {r.get('calls',0):,} | "
                 f"{r.get('directions',0):,} | {r.get('website',0):,} | "
                 f"{r['action_rate']}% | {r['call_rate']}% | {r['mobile_share']}% |")
    L += ["", "---", "", "## סיכום", "",
          f"- **צפיות:** {tot['views']:,}",
          f"- **שיחות טלפון:** {tot['calls']:,}",
          f"- **בקשות ניווט:** {tot['directions']:,}",
          f"- **קליקים לאתר:** {tot['website']:,}",
          f"- **סה\"כ פעולות:** {tot['actions']:,} "
          f"({round(100*tot['actions']/tot['views'],2) if tot['views'] else 0}% מהצפיות)",
          ""]

    weak = [r for r in profiles if r["views"] >= 500 and r["call_rate"] < 2]
    if weak:
        L += ["## פרופילים עם חשיפה ויחס שיחה נמוך", "",
              "יחס מתחת ל-2% אומר שהכרטיס נראה ולא ממיר. בדוק תמונות, ",
              "שעות פעילות, ביקורות ותיאור.", ""]
        for r in weak:
            L.append(f"- **{r.get('name','')[:40]}** — {r['views']:,} צפיות, "
                     f"{r.get('calls',0)} שיחות ({r['call_rate']}%)")
        L.append("")

    L += ["## הערה על מובייל", "",
          "אחוז המובייל קובע מבנה תוכן: פסקאות קצרות, טבלאות שנגללות, ",
          "וכפתור חיוג נגיש. הנתון לכל פרופיל בטבלה למעלה.", ""]
    if failed:
        L += ["## קבצים שלא נותחו", ""] + [f"- `{n}` — {e}" for n, e in failed]

    OUT.mkdir(parents=True, exist_ok=True)
    (OUT / "gbp_performance.md").write_text("\n".join(L), encoding="utf-8")
    (OUT / "gbp_data.json").write_text(
        json.dumps({"date": str(date.today()), "profiles": profiles},
                   ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"נותחו {len(profiles)} פרופילים | צפיות {tot['views']:,} | "
          f"שיחות {tot['calls']:,} | ניווט {tot['directions']:,}")
    print(f"נכתב: {OUT}/gbp_performance.md")
    return 0


if __name__ == "__main__":
    sys.exit(main())
